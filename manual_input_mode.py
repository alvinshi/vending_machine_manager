from itertools import combinations
import os, errno
import random
import openpyxl
from openpyxl import Workbook

from worksheet import Worksheet
import helper as h
from shipment import Shipment
from box import Box

# CONSTANTS_INPUT
REQUIRED_ARGS = 5


OUPTPUT_FOLDER_PATH = "output"
SHIPMENTS = 200
TRY_THRESHOLD = 50
CAPACITY = 100

# String Definition
BACKTRACK = "backtrack"
GENERATE_REPORT = "generate_report"
GENERATE_BRIEF_EXCEL = "output"

RANDOM_MODE = "RAND"
PARTIAL_RANDOM_MODE = "P_RAND"
NON_RANDOM_MODE = "N_RAND"
TOP_UP_MODE = "TOP_UP"

def n_random_in_l(l, n):
	assert(len(l) >= n)
	random.shuffle(l)
	return l[0:n]


def read_mode(str):
	mode = None
	if (str == RANDOM_MODE):
		mode = RANDOM_MODE
	elif (str == PARTIAL_RANDOM_MODE):
		mode = PARTIAL_RANDOM_MODE
	elif (str == NON_RANDOM_MODE):
		mode = NON_RANDOM_MODE
	elif (str == TOP_UP_MODE):
		mode = TOP_UP_MODE
	return mode

# Parse the manual input
def input_parse(t):
	if t == BACKTRACK:
		return (-1, BACKTRACK, 0, 0, 0, 0)
	elif t == GENERATE_REPORT:
		return (-1, GENERATE_REPORT, 0, 0, 0, 0)
	elif t == GENERATE_BRIEF_EXCEL:
		return (-1, GENERATE_BRIEF_EXCEL, 0, 0, 0, 0)

	l_in = t.split()
	if (len(l_in) != REQUIRED_ARGS):
		print("Invalid Input: Please input six parameters with the descripted format")
		return None
	# Parse the input into the desired format
	try:
		box_index = int(l_in[0])
		name = l_in[1].decode('utf-8')
		unit_cost = float(l_in[2])
		numerator, denominator = l_in[3].split("/")
		numerator = int(numerator)
		denominator = int(denominator)
		allocation_mode = read_mode(l_in[4])

		# Chech input validity
		if allocation_mode == None:
			print("Invalid Input: Unrecognizable Mode")
			return None
		if SHIPMENTS % denominator != 0:
			print("Invalid Input: Ns must be a divider of %d" % SHIPMENTS)
			return None
		if (allocation_mode == NON_RANDOM_MODE):
			if (numerator < denominator) or (numerator % denominator != 0):
				print("Invalid Input: In non-random mode, \
					N must larger or equal Ns and Ns must be a divisor or N")
				return None
		if (allocation_mode == RANDOM_MODE):
			print("Warning: We don't recommend you to use this mode no more")
	except:
		print("Invalid Input: %s"% t)
		return None
	return (box_index, name, unit_cost, numerator, denominator, allocation_mode)

def init():
	# Initialize the output directory
	h.create_dir(OUPTPUT_FOLDER_PATH)

	# Initializie the workbook
	workbook = Workbook(write_only=False)
	worksheets = []
	shipments = []
	for i in xrange(SHIPMENTS):
		worksheets.append(Worksheet(workbook.create_sheet("Shipment%d" % (i + 1), i)))
		shipments.append(Shipment(i))
	return (workbook, worksheets, shipments)

def print_allocation_result(allocation_results):
	print("Allocation Result:")
	for key, value in allocation_results.iteritems():
		print("%s : %d" % (str(key), value))

def allocation_results_update(allocation_results, shipment, n):
	if not shipment in allocation_results:
		allocation_results[shipment] = n
	else:
		allocation_results[shipment] = allocation_results[shipment] + n

def worksheet_update(worksheets, index, new_box):
	args = [new_box.box_index, new_box.name, new_box.unit_cost]
	worksheets[index].write(args)

# The length of new_boxes may be larger than SHIPMENT * box_per_shipment
def non_random_allocation(new_boxes, shipments, worksheets, box_per_shipment, allocation_results):
	new_box_sample = new_boxes[0]
	for shipment in shipments:
		for i in xrange(box_per_shipment):
			new_box = new_boxes.pop()
			assigned = shipment.add(new_box)
			## The box is too full
			if (not assigned):
				return (new_box_sample, allocation_results, 
					"Shipments are too full for non-random allocation, please backtrack")

			# Added successfully
			allocation_results_update(allocation_results, shipment, 1)
			worksheet_update(worksheets, shipment.index, new_box)
	return (new_box_sample, allocation_results, None)

def random_allocation(new_boxes, shipments, worksheets, numerator, denominator, allocation_results):
	print("Warning: This method is deprecated in version 2.0")
	new_box_sample = new_boxes[0]
	for new_box in new_boxes:
		assigned = False
		count = 0
		while (not assigned) and (count < TRY_THRESHOLD):
			count += 1
			index = random.randint(0, SHIPMENTS - 1)
			shipment = shipments[index]
			assigned = shipment.add(new_box)
			if assigned:
				allocation_results_update(allocation_results, shipment, 1)
				worksheet_update(worksheets, shipment.index, new_box)
			if (count == TRY_THRESHOLD):
				return (new_box_sample, allocation_results, 
					"Too many items in the box, please try to assign the item manually, an error may exist")
	return (new_box_sample, allocation_results, None)

# Get all shipments that can take in one new box_sample
def get_potential_shipments(shipments, box_sample):
	rtn = []
	for shipment in shipments:
		if (shipment.can_add(box_sample)):
			rtn.append(shipment)
	return rtn

def partial_random_allocation(new_boxes, shipments, worksheets, numerator, denominator, allocation_results):
	assert(len(new_boxes) == SHIPMENTS / denominator * numerator)
	new_box_sample = new_boxes[0]
	if (numerator >= denominator):
		box_per_shipment = numerator / denominator
		numerator = numerator % denominator
		(new_box, allocation_results, msg) = non_random_allocation(new_boxes, shipments, worksheets, box_per_shipment, allocation_results)
		if (msg != None):
			return (new_box, allocation_results, msg) # Failed to allocate
		print("Phase 1: non_random_allocation done")

	# Start partial_random allocation
	potential_shipments = get_potential_shipments(shipments, new_box_sample)
	chosen_shipments_n = SHIPMENTS / denominator * numerator
	if len(potential_shipments) < chosen_shipments_n:
		return (new_box_sample, allocation_results, 
			"Shipments are too full for this allocation, please backtrack")

	chosen_shipments = n_random_in_l(potential_shipments, chosen_shipments_n)
	for shipment in chosen_shipments:
		new_box = new_boxes.pop()
		assigned = shipment.add(new_box)
		assert(assigned)
		allocation_results_update(allocation_results, shipment, 1)
		worksheet_update(worksheets, shipment.index, new_box)
	return (new_box_sample, allocation_results, None)

def top_up(new_boxes, shipments, worksheets, numerator, denominator, allocation_results):
	index = 0
	new_box_sample = new_boxes[0]
	while len(new_boxes) > 0:
		new_box = new_boxes.pop()
		shipment = shipments[index]
		assigned = shipment.add(new_box)
		if assigned:
			allocation_results_update(allocation_results, shipment, 1)
			worksheet_update(worksheets, shipment.index, new_box)
		else:
			new_boxes.append(new_box)
		index = (index + 1) % SHIPMENTS
	return (new_box_sample, allocation_results, None)



def allocate_boxes(new_boxes, shipments, workbook, worksheets, allocation_mode, numerator, denominator):
	assert(len(shipments) == len(worksheets))
	assert(len(shipments) == SHIPMENTS)
	assert(SHIPMENTS % denominator == 0)

	allocation_results = {}
	result = (None, allocation_results, "Unsupported Allocation Mode")
	# Non random allocation
	if (allocation_mode == NON_RANDOM_MODE):
		result = non_random_allocation(new_boxes, shipments, worksheets, numerator/denominator, allocation_results)
	# Random allocation
	elif (allocation_mode == RANDOM_MODE):
		result = random_allocation(new_boxes, shipments, worksheets, numerator, denominator, allocation_results)
	# Partially random allocation
	elif (allocation_mode == PARTIAL_RANDOM_MODE):
		result = partial_random_allocation(new_boxes, shipments, worksheets, numerator, denominator, allocation_results)
	# Final top up
	elif (allocation_mode == TOP_UP_MODE):
		result = top_up(new_boxes, shipments, worksheets, numerator, denominator, allocation_results)
	workbook.save(OUPTPUT_FOLDER_PATH + '/shipments.xlsx')
	return result

def backtrack(new_box, allocation_results, shipments, workbook, worksheets):
	assert(len(shipments) == len(worksheets))
	assert(len(shipments) == SHIPMENTS)
	if (new_box == None): return
	for shipment, amount in allocation_results.iteritems():
		shipment.remove_recent(new_box, amount)
		worksheets[shipment.index].delete_row(amount)
	workbook.save(OUPTPUT_FOLDER_PATH + '/shipments.xlsx')
	# Destroy the allocation_results after backtracking
	allocation_results.clear()

def generate_report(shipments):
	f = open(OUPTPUT_FOLDER_PATH+"/report.txt", "w+")

	# Comparator
	def get_index(shipment):
		return shipment.index

	shipments.sort(key=get_index)
	for shipment in shipments:
		f.write("Shipment %d :\n" % (shipment.index + 1))
		length = len(shipment.boxes)
		f.write("  Total Boxes: %d\n" % length)
		f.write("  Total Cost:  {:.2f}\n".format(shipment.total_cost))
		if (length > 0):
			f.write("  Average Cost:  {:.2f}\n".format(shipment.total_cost / len(shipment.boxes)))
			f.write("  Big Price Amount (>= 200): %d\n" % shipment.big_prize_threshold_counter)
			f.write("  Big Price Amount (>= 30): %d\n" % shipment.prize_threshold_counter)
	f.close()

def generate_brief_excel(shipments):
	# Initializie the workbook
	workbook = Workbook(write_only=False)
	worksheets = []
	for shipment in shipments:
		i = shipment.index
		worksheet = Worksheet(workbook.create_sheet("Shipment%d" % (i+ 1), i))
		for box in shipment.boxes:
			worksheet.write([box.box_index, box.name])
	workbook.save(OUPTPUT_FOLDER_PATH + '/shipments_brief.xlsx')

# Serve as the main message dispatcher
def manual_input_mode():
	is_full = False

	h.print_sub_title("MANUAL INPUT MODE")
	print("Keywords")
	print("RAND: Random Allocation Mode")
	print("P_RAND: Partially Random Allocation Mode")
	print("N_RAND: None Random Allocation Mode")
	print("TOP_UP: Top Up Mode")
	(workbook, worksheets, shipments) = init()

	# Start to take in inputs
	new_box = None
	allocation_results = {}
	total_boxes = 0
	while (total_boxes < SHIPMENTS * CAPACITY):
		print("")
		print("Please enter '<Product_Index> <Product_Name> <Product_Price> <N/Ns> <RAND/P_RAND/N_RAND/TOP_UP>'")
		print("or 'backtrack' or 'generate_report' or 'output' or 'q': ")
		text_in = h.Raw_Input()
		parsed = input_parse(text_in) # it returns None if parsing failed
		if (parsed):
			box_index, name, unit_cost, numerator, denominator, allocation_mode = parsed
			if name == BACKTRACK:
				backtrack(new_box, allocation_results, shipments, workbook, worksheets)
			elif name == GENERATE_REPORT:
				generate_report(shipments)
			elif name == GENERATE_BRIEF_EXCEL:
				generate_brief_excel(shipments)
			else:
				# Create new box objects
				new_box_total = SHIPMENTS / denominator * numerator
				if (total_boxes + new_box_total) > SHIPMENTS * CAPACITY:
					print("Error: Not enough space for %d new boxes" % new_box_total)
					print("You already have %d boxes in the shipments, %d more required" % (total_boxes, SHIPMENTS * CAPACITY - total_boxes))
				else:
					new_boxes = []
					for i in xrange(SHIPMENTS / denominator * numerator):
						new_boxes.append(Box(box_index, name, unit_cost, numerator, denominator))
					# Allocate new_boxes
					(new_box, allocation_results, error) = \
					allocate_boxes(new_boxes, shipments, workbook, worksheets, allocation_mode, numerator, denominator)
					if (error != None):
						print(error)
					else: 
						total_boxes += SHIPMENTS / denominator * numerator
						print("Allocation Succeeded, %d boxes are allocated" % new_box_total)
